"""
Tests for src/data_loader.py.

Run from the project root:

    python tests/test_data_loader.py

Covers a real crash found in testing: a spreadsheet with two columns that
collapse to the same name after cleaning made every later groupby() on that
name raise `ValueError: cannot insert X, already exists`. See
_deduplicate_columns() in data_loader.py for the fix.
"""

import io
import os
import sys

import openpyxl
import pandas as pd

sys.path.insert(0, os.path.dirname(os.path.dirname(os.path.abspath(__file__))))

from src.data_loader import (  # noqa: E402
    SHEET_COLUMN,
    clean_dataframe,
    load_combined,
    load_dataframe,
)
from src.analyzer import ColumnMapping, analyze_budget, detect_columns  # noqa: E402


def _xlsx_bytes(headers, rows) -> bytes:
    """Build a real .xlsx file in memory, so tests exercise the actual
    openpyxl round-trip rather than a DataFrame constructed by hand."""
    workbook = openpyxl.Workbook()
    sheet = workbook.active
    sheet.append(headers)
    for row in rows:
        sheet.append(row)
    buffer = io.BytesIO()
    workbook.save(buffer)
    return buffer.getvalue()


def test_clean_dataframe_deduplicates_identical_column_names():
    df = pd.DataFrame([["A", 1], ["B", 2]])
    df.columns = ["Department", "Department"]   # assigned directly — pandas'
    # own read_excel would never hand us this, but a manually built or
    # concatenated DataFrame could, and the fix must not assume the source.

    cleaned = clean_dataframe(df)

    assert list(cleaned.columns) == ["Department", "Department (2)"]
    assert len(set(cleaned.columns)) == len(cleaned.columns)


def test_clean_dataframe_deduplicates_names_that_collide_after_whitespace_cleanup():
    """
    The real-world trigger: a header with a trailing space.

    "Department" and "Department " read from Excel as two distinct strings —
    pandas has no reason to rename either — but clean_dataframe's own
    whitespace normalisation collapses them to the same name. Without
    deduplication, that normalisation step would be the thing creating the
    crash it was meant to prevent.
    """
    df = pd.DataFrame([["A", 1], ["B", 2]])
    df.columns = ["Department", "Department "]

    cleaned = clean_dataframe(df)

    assert list(cleaned.columns) == ["Department", "Department (2)"]


def test_deduplication_survives_three_or_more_collisions():
    df = pd.DataFrame([[1, 2, 3, 4]])
    df.columns = ["Cost", "Cost", "Cost", "Cost"]

    cleaned = clean_dataframe(df)

    assert list(cleaned.columns) == ["Cost", "Cost (2)", "Cost (3)", "Cost (4)"]


def test_loading_a_real_workbook_with_a_duplicate_header_does_not_crash():
    """
    End-to-end: the exact failure reported in testing.

    A department-analysis column and an unrelated free-text column share a
    header (once whitespace is normalised); loading, mapping, and grouping the
    file must all complete rather than raising downstream in analyzer.py.
    """
    file_bytes = _xlsx_bytes(
        headers=["Department", "Budget", "Actual", "Department "],
        rows=[
            ["Sales", 100, 110, "Sales-East"],
            ["Sales", 200, 190, "Sales-West"],
            ["IT", 300, 330, "IT-Core"],
        ],
    )

    df = load_dataframe(file_bytes, "budget.xlsx")
    assert len(set(df.columns)) == len(df.columns), "Duplicate columns survived loading"

    mapping = ColumnMapping(budget="Budget", actual="Actual", department="Department")
    analysis = analyze_budget(df, mapping)

    # This is the exact call that crashed: grouping by a column name that,
    # before the fix, was no longer unique in the DataFrame.
    grouped = analysis.by_group("Department")
    assert set(grouped["Department"]) == {"Sales", "IT"}

    ordered = analysis.by_group_ordered("Department")
    assert set(ordered["Department"]) == {"Sales", "IT"}


def test_column_detection_still_works_alongside_a_renamed_duplicate():
    """The renamed column must not confuse auto-detection of the real ones."""
    file_bytes = _xlsx_bytes(
        headers=["Department", "Budgeted Amount", "Actual Amount", "Department "],
        rows=[["Sales", 100, 110, "Note"]],
    )

    df = load_dataframe(file_bytes, "budget.xlsx")
    mapping = detect_columns(df)

    assert mapping.budget == "Budgeted Amount"
    assert mapping.actual == "Actual Amount"
    assert mapping.department == "Department"


# ---------------------------------------------------------------------------
# Header-row and label detection, from a real submitted workbook
# ---------------------------------------------------------------------------
def test_junk_column_headers_lose_to_the_real_header_row():
    """
    Excel writes "Column1, Column2, ..." above a table with no header styling.

    Those are not blank, so nothing flags them as placeholders — and because
    column detection falls back to picking the first two *numeric* columns
    when headers say nothing, the junk row still yields a complete mapping.
    It has to lose to the real headers on some other signal, or the whole
    report is built on columns named Column2 and Column3.
    """
    from app import detect_header_row   # imported here to keep Streamlit out
    # of this module's import path for the other tests.

    file_bytes = _xlsx_bytes(
        headers=["Column1", "Column2", "Column3", "Column4"],
        rows=[
            ["Quarter", "Total Budget", "Total Actual", "Variance"],
            ["Q1", 749000, 758100, 9100],
            ["Q2", 786450, 796005, 9555],
        ],
    )

    assert detect_header_row(file_bytes, "budget.xlsx", None) == 1

    df = load_dataframe(file_bytes, "budget.xlsx", header_row=1)
    mapping = detect_columns(df)
    assert mapping.budget == "Total Budget"
    assert mapping.actual == "Total Actual"
    assert mapping.period == "Quarter"


def test_a_misspelled_label_header_still_yields_a_grouping_column():
    """
    A real workbook labelled its department column "Deparment annual
    spending" — misspelled, and unlike any keyword we know.

    Without a fallback the file would load with no label column at all: no
    grouping, no charts, no drill-down, just one grand total. Falling back to
    the first text column is far better than that.
    """
    file_bytes = _xlsx_bytes(
        headers=["Deparment annual spending", "Total Budget", "Total Actual"],
        rows=[["Sales", 2150000, 2193000], ["HR", 516000, 537500]],
    )

    df = load_dataframe(file_bytes, "budget.xlsx")
    mapping = detect_columns(df)

    assert mapping.budget == "Total Budget"
    assert mapping.actual == "Total Actual"
    assert mapping.grouping_columns() == ["Deparment annual spending"], \
        "No label column detected — the report would have nothing to group by"


def test_label_fallback_does_not_override_a_properly_detected_column():
    """The fallback is a last resort, not a competitor to real detection."""
    file_bytes = _xlsx_bytes(
        headers=["Notes", "Department", "Budget", "Actual"],
        rows=[["ignore me", "Sales", 100, 110]],
    )

    df = load_dataframe(file_bytes, "budget.xlsx")
    mapping = detect_columns(df)

    assert mapping.department == "Department"
    assert "Notes" not in mapping.grouping_columns()


# ---------------------------------------------------------------------------
# Combining a sheet-per-quarter workbook
# ---------------------------------------------------------------------------
def _quarterly_workbook() -> bytes:
    """A workbook shaped like a real one: four quarter tabs plus two summary
    tabs with different columns that must NOT be swept into the combination."""
    workbook = openpyxl.Workbook()

    summary = workbook.active
    summary.title = "Actual vs Budget"
    summary.append(["Quarter", "Total Budget", "Total Actual"])
    summary.append(["Q1", 749000, 758100])

    for index, quarter in enumerate(["Q1", "Q2", "Q3", "Q4"], start=1):
        sheet = workbook.create_sheet(quarter)
        sheet.append(["Category", "Department", "Budget", "Actual"])
        sheet.append(["Revenue", "Sales", 100 * index, 110 * index])
        sheet.append(["Payroll", "HR", 200 * index, 190 * index])

    buffer = io.BytesIO()
    workbook.save(buffer)
    return buffer.getvalue()


def test_combining_stacks_matching_sheets_and_records_their_names():
    file_bytes = _quarterly_workbook()
    pairs = [(q, 0) for q in ["Q1", "Q2", "Q3", "Q4"]]

    combined = load_combined(file_bytes, "budget.xlsx", pairs)

    assert len(combined) == 8                      # 2 rows × 4 quarters
    assert list(combined.columns)[0] == SHEET_COLUMN
    assert list(combined[SHEET_COLUMN].unique()) == ["Q1", "Q2", "Q3", "Q4"]
    # Row order follows the sheet order given, not alphabetical or arbitrary.
    assert combined.iloc[0][SHEET_COLUMN] == "Q1"
    assert combined[SHEET_COLUMN].tolist().count("Q3") == 2


def test_combined_sheets_analyse_as_one_budget():
    """
    The point of combining: a department's whole year in one analysis, with
    the quarter available as a period column for the trend charts.
    """
    file_bytes = _quarterly_workbook()
    pairs = [(q, 0) for q in ["Q1", "Q2", "Q3", "Q4"]]

    combined = load_combined(file_bytes, "budget.xlsx", pairs)
    mapping = detect_columns(combined)

    assert mapping.budget == "Budget"
    assert mapping.actual == "Actual"
    assert mapping.department == "Department"
    # "Period" is recognised by name, so the quarter charts work with no
    # further configuration — that is why the column is called Period.
    assert mapping.period == SHEET_COLUMN

    analysis = analyze_budget(combined, mapping)
    # Sales across four quarters: 100+200+300+400 budgeted.
    totals = analysis.totals_for("Department", "Sales")
    assert totals["budget"] == 1000
    assert totals["actual"] == 1100
    assert totals["line_items"] == 4

    quarters = analysis.by_group_ordered(SHEET_COLUMN)
    assert list(quarters[SHEET_COLUMN]) == ["Q1", "Q2", "Q3", "Q4"]


def test_combining_skips_an_empty_sheet_rather_than_failing():
    """One blank tab must not sink the whole combination."""
    workbook = openpyxl.Workbook()
    first = workbook.active
    first.title = "Q1"
    first.append(["Department", "Budget", "Actual"])
    first.append(["Sales", 100, 110])
    workbook.create_sheet("Q2")          # deliberately empty

    buffer = io.BytesIO()
    workbook.save(buffer)

    combined = load_combined(buffer.getvalue(), "b.xlsx", [("Q1", 0), ("Q2", 0)])
    assert len(combined) == 1
    assert list(combined[SHEET_COLUMN].unique()) == ["Q1"]


def test_period_column_name_avoids_colliding_with_an_existing_one():
    """A file that already has a "Period" column must not end up with two."""
    workbook = openpyxl.Workbook()
    for quarter in ["Q1", "Q2"]:
        sheet = workbook.active if quarter == "Q1" else workbook.create_sheet()
        sheet.title = quarter
        sheet.append(["Period", "Department", "Budget", "Actual"])
        sheet.append(["month 1", "Sales", 100, 110])

    buffer = io.BytesIO()
    workbook.save(buffer)

    combined = load_combined(buffer.getvalue(), "b.xlsx", [("Q1", 0), ("Q2", 0)])

    assert len(set(combined.columns)) == len(combined.columns)
    assert "Period" in combined.columns and "Period (2)" in combined.columns


def test_source_variance_column_is_disclosed_not_silently_ignored():
    """
    Spreadsheets disagree on which way round variance is subtracted. When the
    file brings its own column we ignore it — and must say so, or a reader
    comparing the two will think one of them is a bug.
    """
    file_bytes = _xlsx_bytes(
        headers=["Department", "Budget", "Actual", "Variance"],
        rows=[["Sales", 100, 110, -10]],      # their sign convention: budget - actual
    )

    df = load_dataframe(file_bytes, "budget.xlsx")
    analysis = analyze_budget(df, detect_columns(df))

    assert analysis.variance_column == "Variance (calculated)"
    assert analysis.data["Variance (calculated)"].iloc[0] == 10     # actual - budget
    assert any("already has a 'Variance' column" in w for w in analysis.warnings), \
        analysis.warnings


def main():
    tests = [v for k, v in sorted(globals().items()) if k.startswith("test_")]
    failures = 0

    for test in tests:
        try:
            test()
            print("  PASS  {}".format(test.__name__))
        except Exception as error:  # noqa: BLE001 - report and continue
            failures += 1
            print("  FAIL  {}: {}".format(test.__name__, error))

    print("\n{} passed, {} failed".format(len(tests) - failures, failures))
    return 1 if failures else 0


if __name__ == "__main__":
    sys.exit(main())
